"""数据点管理API"""
import logging
from io import BytesIO
from typing import Optional, List

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, or_
from openpyxl import load_workbook

from app.database import get_db
from app.models.data_point import DataPoint, DataItem
from app.models.device import Device
from app.models.user import User
from app.schemas import DataPointCreate, DataPointUpdate, DataItemCreate, ResponseModel
from app.core.security import get_current_user, require_admin

logger = logging.getLogger("industrial-monitor")

router = APIRouter(prefix="/data-points", tags=["数据点管理"])


# ============ 数据点列表 ============

@router.get("", response_model=ResponseModel)
async def list_data_points(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    keyword: Optional[str] = None,
    hierarchy_id: Optional[int] = None,
    device_id: Optional[int] = None,
    status: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """数据点列表"""
    query = select(DataPoint)
    count_query = select(func.count(DataPoint.id))

    if keyword:
        like_pattern = f"%{keyword}%"
        filter_cond = or_(
            DataPoint.point_code.ilike(like_pattern),
            DataPoint.point_name.ilike(like_pattern),
        )
        query = query.where(filter_cond)
        count_query = count_query.where(filter_cond)
    if hierarchy_id is not None:
        query = query.where(DataPoint.hierarchy_id == hierarchy_id)
        count_query = count_query.where(DataPoint.hierarchy_id == hierarchy_id)
    if device_id is not None:
        query = query.where(DataPoint.device_id == device_id)
        count_query = count_query.where(DataPoint.device_id == device_id)
    if status:
        query = query.where(DataPoint.status == status)
        count_query = count_query.where(DataPoint.status == status)

    total = (await db.execute(count_query)).scalar()
    query = query.order_by(DataPoint.id.desc()).offset((page - 1) * page_size).limit(page_size)
    result = await db.execute(query)
    points = result.scalars().all()

    return ResponseModel(data={
        "list": [
            {
                "id": p.id,
                "point_code": p.point_code,
                "point_name": p.point_name,
                "hierarchy_id": p.hierarchy_id,
                "device_id": p.device_id,
                "location_desc": p.location_desc,
                "longitude": p.longitude,
                "latitude": p.latitude,
                "status": p.status,
                "item_count": len(p.items) if p.items else 0,
                "created_at": p.created_at.isoformat() if p.created_at else None,
            }
            for p in points
        ],
        "total": total,
        "page": page,
        "page_size": page_size,
    })


# ============ 数据点详情 ============

@router.get("/{point_id}", response_model=ResponseModel)
async def get_data_point(
    point_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """数据点详情（含数据项）"""
    point = await db.get(DataPoint, point_id)
    if not point:
        raise HTTPException(status_code=404, detail="数据点不存在")

    return ResponseModel(data={
        "id": point.id,
        "point_code": point.point_code,
        "point_name": point.point_name,
        "hierarchy_id": point.hierarchy_id,
        "device_id": point.device_id,
        "location_desc": point.location_desc,
        "longitude": point.longitude,
        "latitude": point.latitude,
        "status": point.status,
        "created_at": point.created_at.isoformat() if point.created_at else None,
        "updated_at": point.updated_at.isoformat() if point.updated_at else None,
        "items": [
            {
                "id": item.id,
                "item_code": item.item_code,
                "item_name": item.item_name,
                "unit": item.unit,
                "data_type": item.data_type,
                "min_value": item.min_value,
                "max_value": item.max_value,
                "alarm_enabled": item.alarm_enabled,
                "alarm_low": item.alarm_low,
                "alarm_high": item.alarm_high,
                "alarm_level": item.alarm_level,
                "scale": item.scale,
                "offset": item.offset,
                "sort_order": item.sort_order,
            }
            for item in (point.items or [])
        ],
    })


# ============ 新增数据点 ============

@router.post("", response_model=ResponseModel)
async def create_data_point(
    req: DataPointCreate,
    current_user: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """新增数据点（可同时创建数据项）"""
    # 检查编码唯一
    existing = await db.execute(
        select(DataPoint).where(DataPoint.point_code == req.point_code)
    )
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="数据点编码已存在")

    point = DataPoint(
        point_code=req.point_code,
        point_name=req.point_name,
        hierarchy_id=req.hierarchy_id,
        device_id=req.device_id,
        location_desc=req.location_desc,
        longitude=req.longitude,
        latitude=req.latitude,
    )
    db.add(point)
    await db.flush()

    # 创建关联数据项
    for idx, item_req in enumerate(req.items):
        item = DataItem(
            point_id=point.id,
            item_code=item_req.item_code,
            item_name=item_req.item_name,
            unit=item_req.unit,
            data_type=item_req.data_type,
            min_value=item_req.min_value,
            max_value=item_req.max_value,
            alarm_enabled=item_req.alarm_enabled,
            alarm_low=item_req.alarm_low,
            alarm_high=item_req.alarm_high,
            alarm_level=item_req.alarm_level,
            scale=item_req.scale,
            offset=item_req.offset,
            sort_order=idx,
        )
        db.add(item)

    await db.commit()
    await db.refresh(point)
    logger.info(f"新增数据点: {point.point_code} (ID={point.id})")
    return ResponseModel(data={"id": point.id}, message="创建成功")


# ============ 修改数据点 ============

@router.put("/{point_id}", response_model=ResponseModel)
async def update_data_point(
    point_id: int,
    req: DataPointUpdate,
    current_user: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """修改数据点"""
    point = await db.get(DataPoint, point_id)
    if not point:
        raise HTTPException(status_code=404, detail="数据点不存在")

    update_data = req.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(point, key, value)

    await db.commit()
    logger.info(f"修改数据点: ID={point_id}")
    return ResponseModel(message="修改成功")


# ============ 删除数据点 ============

@router.delete("/{point_id}", response_model=ResponseModel)
async def delete_data_point(
    point_id: int,
    current_user: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """删除数据点（级联删除数据项）"""
    point = await db.get(DataPoint, point_id)
    if not point:
        raise HTTPException(status_code=404, detail="数据点不存在")

    await db.delete(point)
    await db.commit()
    logger.info(f"删除数据点: ID={point_id}")
    return ResponseModel(message="删除成功")


# ============ 数据项 CRUD ============

@router.get("/{point_id}/items", response_model=ResponseModel)
async def list_data_items(
    point_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """获取数据点下的数据项列表"""
    point = await db.get(DataPoint, point_id)
    if not point:
        raise HTTPException(status_code=404, detail="数据点不存在")

    result = await db.execute(
        select(DataItem)
        .where(DataItem.point_id == point_id)
        .order_by(DataItem.sort_order)
    )
    items = result.scalars().all()

    return ResponseModel(data=[
        {
            "id": item.id,
            "point_id": item.point_id,
            "item_code": item.item_code,
            "item_name": item.item_name,
            "unit": item.unit,
            "data_type": item.data_type,
            "min_value": item.min_value,
            "max_value": item.max_value,
            "alarm_enabled": item.alarm_enabled,
            "alarm_low": item.alarm_low,
            "alarm_high": item.alarm_high,
            "alarm_level": item.alarm_level,
            "validation_enabled": item.validation_enabled,
            "jump_threshold": item.jump_threshold,
            "jump_window_size": item.jump_window_size,
            "zero_count_limit": item.zero_count_limit,
            "dup_count_limit": item.dup_count_limit,
            "scale": item.scale,
            "offset": item.offset,
            "sort_order": item.sort_order,
            "created_at": item.created_at.isoformat() if item.created_at else None,
        }
        for item in items
    ])


@router.post("/{point_id}/items", response_model=ResponseModel)
async def create_data_item(
    point_id: int,
    req: DataItemCreate,
    current_user: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """新增数据项"""
    point = await db.get(DataPoint, point_id)
    if not point:
        raise HTTPException(status_code=404, detail="数据点不存在")

    # 检查同一数据点下编码唯一
    existing = await db.execute(
        select(DataItem).where(
            DataItem.point_id == point_id,
            DataItem.item_code == req.item_code,
        )
    )
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="该数据点下已存在相同编码的数据项")

    # 获取当前最大排序号
    max_sort = await db.execute(
        select(func.max(DataItem.sort_order)).where(DataItem.point_id == point_id)
    )
    next_sort = (max_sort.scalar() or 0) + 1

    item = DataItem(
        point_id=point_id,
        item_code=req.item_code,
        item_name=req.item_name,
        unit=req.unit,
        data_type=req.data_type,
        min_value=req.min_value,
        max_value=req.max_value,
        alarm_enabled=req.alarm_enabled,
        alarm_low=req.alarm_low,
        alarm_high=req.alarm_high,
        alarm_level=req.alarm_level,
        scale=req.scale,
        offset=req.offset,
        sort_order=next_sort,
    )
    db.add(item)
    await db.flush()
    await db.refresh(item)

    return ResponseModel(data={"id": item.id}, message="创建成功")


@router.put("/{point_id}/items/{item_id}", response_model=ResponseModel)
async def update_data_item(
    point_id: int,
    item_id: int,
    req: dict,
    current_user: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """修改数据项"""
    item = await db.get(DataItem, item_id)
    if not item or item.point_id != point_id:
        raise HTTPException(status_code=404, detail="数据项不存在")

    for key, value in req.items():
        if hasattr(item, key) and key not in ("id", "point_id", "created_at"):
            setattr(item, key, value)

    await db.commit()
    return ResponseModel(message="修改成功")


@router.delete("/{point_id}/items/{item_id}", response_model=ResponseModel)
async def delete_data_item(
    point_id: int,
    item_id: int,
    current_user: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """删除数据项"""
    item = await db.get(DataItem, item_id)
    if not item or item.point_id != point_id:
        raise HTTPException(status_code=404, detail="数据项不存在")

    await db.delete(item)
    await db.commit()
    return ResponseModel(message="删除成功")


# ============ 批量导入 ============

@router.post("/import", response_model=ResponseModel)
async def import_data_points(
    file: UploadFile = File(...),
    current_user: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """从Excel批量导入数据点"""
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="请上传Excel文件(.xlsx)")

    content = await file.read()
    wb = load_workbook(BytesIO(content), read_only=True)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    header_map = {h.strip(): idx for idx, h in enumerate(headers) if h}

    required_fields = ["数据点编码", "数据点名称", "所属层级ID"]
    for field in required_fields:
        if field not in header_map:
            raise HTTPException(status_code=400, detail=f"缺少必填列: {field}")

    imported = 0
    skipped = 0
    errors = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            point_code = str(row[header_map["数据点编码"]]).strip() if row[header_map["数据点编码"]] else None
            if not point_code:
                continue

            existing = await db.execute(
                select(DataPoint).where(DataPoint.point_code == point_code)
            )
            if existing.scalar_one_or_none():
                skipped += 1
                continue

            point = DataPoint(
                point_code=point_code,
                point_name=str(row[header_map["数据点名称"]]).strip(),
                hierarchy_id=int(row[header_map["所属层级ID"]]),
                device_id=int(row[header_map["设备ID"]]) if "设备ID" in header_map and row[header_map["设备ID"]] else None,
                location_desc=str(row[header_map.get("位置描述", -1)]).strip() if "位置描述" in header_map and row[header_map["位置描述"]] else None,
            )
            db.add(point)
            imported += 1
        except Exception as e:
            errors.append(f"第{row_idx}行: {str(e)}")

    await db.commit()
    wb.close()

    return ResponseModel(data={
        "imported": imported,
        "skipped": skipped,
        "errors": errors[:50],
    }, message=f"导入完成: 新增 {imported}, 跳过 {skipped}")
