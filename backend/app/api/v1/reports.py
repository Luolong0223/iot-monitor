"""报表API"""
import logging
from io import BytesIO
from datetime import datetime, timedelta, timezone
from typing import Optional, List

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from openpyxl import Workbook

from app.database import get_db
from app.models.data_point import DataPoint, DataItem
from app.models.device import Device
from app.models.alarm import AlarmRecord
from app.models.hierarchy import HierarchyLevel
from app.models.user import User
from app.schemas import ResponseModel
from app.core.security import get_current_user
from app.services.tdengine_service import get_tdengine_service

logger = logging.getLogger("industrial-monitor")

router = APIRouter(prefix="/reports", tags=["报表管理"])


@router.get("/daily", response_model=ResponseModel)
async def daily_report(
    date: Optional[str] = Query(None, description="日期 YYYY-MM-DD，默认昨天"),
    hierarchy_id: Optional[int] = None,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """日报表"""
    if date:
        try:
            report_date = datetime.strptime(date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
        except ValueError:
            raise HTTPException(status_code=400, detail="日期格式错误，应为 YYYY-MM-DD")
    else:
        report_date = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0) - timedelta(days=1)

    start_time = report_date
    end_time = report_date + timedelta(days=1)

    return await _generate_report(db, start_time, end_time, hierarchy_id, "日报表")


@router.get("/monthly", response_model=ResponseModel)
async def monthly_report(
    year: int = Query(..., ge=2020, le=2100),
    month: int = Query(..., ge=1, le=12),
    hierarchy_id: Optional[int] = None,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """月报表"""
    start_time = datetime(year, month, 1, tzinfo=timezone.utc)
    if month == 12:
        end_time = datetime(year + 1, 1, 1, tzinfo=timezone.utc)
    else:
        end_time = datetime(year, month + 1, 1, tzinfo=timezone.utc)

    return await _generate_report(db, start_time, end_time, hierarchy_id, f"{year}年{month}月报表")


@router.post("/custom", response_model=ResponseModel)
async def custom_report(
    start_time: str = Query(..., description="开始时间 ISO格式"),
    end_time: str = Query(..., description="结束时间 ISO格式"),
    hierarchy_id: Optional[int] = None,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """自定义时间范围报表"""
    start = _parse_dt(start_time)
    end = _parse_dt(end_time)

    if start >= end:
        raise HTTPException(status_code=400, detail="开始时间必须早于结束时间")

    return await _generate_report(db, start, end, hierarchy_id, "自定义报表")


@router.get("/export")
async def export_report(
    report_type: str = Query("daily", description="报表类型: daily/monthly/custom"),
    start_time: Optional[str] = None,
    end_time: Optional[str] = None,
    year: Optional[int] = None,
    month: Optional[int] = None,
    date: Optional[str] = None,
    hierarchy_id: Optional[int] = None,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """导出报表到Excel"""
    # 确定时间范围
    now = datetime.now(timezone.utc)

    if report_type == "daily":
        if date:
            report_date = datetime.strptime(date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
        else:
            report_date = now.replace(hour=0, minute=0, second=0, microsecond=0) - timedelta(days=1)
        start = report_date
        end = report_date + timedelta(days=1)
        sheet_name = "日报表"
    elif report_type == "monthly":
        y = year or now.year
        m = month or now.month
        start = datetime(y, m, 1, tzinfo=timezone.utc)
        end = datetime(y + (1 if m == 12 else 0), 1 if m == 12 else m + 1, 1, tzinfo=timezone.utc)
        sheet_name = f"{y}年{m}月"
    else:
        start = _parse_dt(start_time or (now - timedelta(days=7)).isoformat())
        end = _parse_dt(end_time or now.isoformat())
        sheet_name = "自定义报表"

    # 查询设备统计
    total_devices = (await db.execute(select(func.count(Device.id)))).scalar() or 0
    online_devices = (await db.execute(
        select(func.count(Device.id)).where(Device.status == "online")
    )).scalar() or 0

    # 查询告警统计
    alarm_count = (await db.execute(
        select(func.count(AlarmRecord.id)).where(
            AlarmRecord.created_at >= start, AlarmRecord.created_at < end
        )
    )).scalar() or 0

    # 获取数据点统计
    tdengine = get_tdengine_service()
    points_result = await db.execute(select(DataPoint).where(DataPoint.status == "active"))
    points = points_result.scalars().all()

    wb = Workbook()

    # 概览页
    ws = wb.active
    ws.title = "概览"
    ws.append(["报表类型", sheet_name])
    ws.append(["时间范围", f"{start.isoformat()} ~ {end.isoformat()}"])
    ws.append(["设备总数", total_devices])
    ws.append(["在线设备", online_devices])
    ws.append(["告警次数", alarm_count])
    ws.append([])

    # 数据统计页
    ws2 = wb.create_sheet("数据统计")
    ws2.append(["数据点", "数据项", "单位", "平均值", "最大值", "最小值", "数据条数"])

    for point in points[:100]:  # 限制100个数据点
        for item in (point.items or []):
            stats = None
            if tdengine.connected:
                stats = tdengine.query_statistics(point.id, item.id, start, end)
            ws2.append([
                point.point_name,
                item.item_name,
                item.unit or "",
                stats.get("avg") if stats else "",
                stats.get("max") if stats else "",
                stats.get("min") if stats else "",
                stats.get("count") if stats else 0,
            ])

    # 告警明细页
    ws3 = wb.create_sheet("告警明细")
    ws3.append(["时间", "数据点ID", "数据项ID", "类型", "级别", "值", "阈值", "状态"])

    alarms_result = await db.execute(
        select(AlarmRecord)
        .where(AlarmRecord.created_at >= start, AlarmRecord.created_at < end)
        .order_by(AlarmRecord.created_at.desc())
        .limit(5000)
    )
    for a in alarms_result.scalars():
        ws3.append([
            a.created_at.isoformat() if a.created_at else "",
            a.point_id,
            a.item_id,
            a.alarm_type,
            a.alarm_level,
            a.alarm_value,
            a.threshold,
            a.status,
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    wb.close()

    filename = f"report_{report_type}_{start.strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


async def _generate_report(
    db: AsyncSession,
    start_time: datetime,
    end_time: datetime,
    hierarchy_id: Optional[int],
    title: str,
) -> ResponseModel:
    """生成报表数据"""
    # 设备统计
    total_devices = (await db.execute(select(func.count(Device.id)))).scalar() or 0
    online_devices = (await db.execute(
        select(func.count(Device.id)).where(Device.status == "online")
    )).scalar() or 0

    # 告警统计
    alarm_count = (await db.execute(
        select(func.count(AlarmRecord.id)).where(
            AlarmRecord.created_at >= start_time,
            AlarmRecord.created_at < end_time,
        )
    )).scalar() or 0

    critical_count = (await db.execute(
        select(func.count(AlarmRecord.id)).where(
            AlarmRecord.created_at >= start_time,
            AlarmRecord.created_at < end_time,
            AlarmRecord.alarm_level == "critical",
        )
    )).scalar() or 0

    # 数据点统计
    total_points = (await db.execute(select(func.count(DataPoint.id)))).scalar() or 0

    return ResponseModel(data={
        "title": title,
        "time_range": {
            "start": start_time.isoformat(),
            "end": end_time.isoformat(),
        },
        "devices": {
            "total": total_devices,
            "online": online_devices,
            "online_rate": round(online_devices / total_devices * 100, 2) if total_devices else 0,
        },
        "data_points": {
            "total": total_points,
        },
        "alarms": {
            "total": alarm_count,
            "critical": critical_count,
        },
    })


def _parse_dt(dt_str: str) -> datetime:
    """解析日期时间"""
    try:
        if dt_str.endswith("Z"):
            dt_str = dt_str[:-1] + "+00:00"
        return datetime.fromisoformat(dt_str)
    except ValueError:
        raise HTTPException(status_code=400, detail=f"时间格式错误: {dt_str}")
