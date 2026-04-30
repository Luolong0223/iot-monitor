"""设备管理API"""
import logging
from io import BytesIO
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, or_
from openpyxl import Workbook, load_workbook

from app.database import get_db
from app.models.device import Device
from app.models.data_point import DataPoint
from app.models.user import User
from app.schemas import DeviceCreate, DeviceUpdate, DeviceResponse, ResponseModel
from app.core.security import get_current_user, require_admin

logger = logging.getLogger("industrial-monitor")

router = APIRouter(prefix="/devices", tags=["设备管理"])


# ============ 列表查询 ============

@router.get("", response_model=ResponseModel)
async def list_devices(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    keyword: Optional[str] = None,
    hierarchy_id: Optional[int] = None,
    device_type: Optional[str] = None,
    status: Optional[str] = None,
    comm_mode: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """设备列表（支持分页、筛选、搜索）"""
    query = select(Device)
    count_query = select(func.count(Device.id))

    # 关键字搜索
    if keyword:
        like_pattern = f"%{keyword}%"
        filter_cond = or_(
            Device.device_code.ilike(like_pattern),
            Device.device_name.ilike(like_pattern),
            Device.sim_card.ilike(like_pattern),
        )
        query = query.where(filter_cond)
        count_query = count_query.where(filter_cond)

    # 筛选条件
    if hierarchy_id is not None:
        query = query.where(Device.hierarchy_id == hierarchy_id)
        count_query = count_query.where(Device.hierarchy_id == hierarchy_id)
    if device_type:
        query = query.where(Device.device_type == device_type)
        count_query = count_query.where(Device.device_type == device_type)
    if status:
        query = query.where(Device.status == status)
        count_query = count_query.where(Device.status == status)
    if comm_mode:
        query = query.where(Device.comm_mode == comm_mode)
        count_query = count_query.where(Device.comm_mode == comm_mode)

    total = (await db.execute(count_query)).scalar()
    query = query.order_by(Device.id.desc()).offset((page - 1) * page_size).limit(page_size)
    result = await db.execute(query)
    devices = result.scalars().all()

    return ResponseModel(data={
        "list": [
            {
                "id": d.id,
                "device_code": d.device_code,
                "device_name": d.device_name,
                "device_type": d.device_type,
                "hierarchy_id": d.hierarchy_id,
                "comm_mode": d.comm_mode,
                "protocol_id": d.protocol_id,
                "sim_card": d.sim_card,
                "install_date": str(d.install_date) if d.install_date else None,
                "status": d.status,
                "last_heartbeat": d.last_heartbeat.isoformat() if d.last_heartbeat else None,
                "firmware_version": d.firmware_version,
                "remark": d.remark,
                "created_at": d.created_at.isoformat() if d.created_at else None,
            }
            for d in devices
        ],
        "total": total,
        "page": page,
        "page_size": page_size,
    })


# ============ 新增设备 ============

@router.post("", response_model=ResponseModel)
async def create_device(
    req: DeviceCreate,
    current_user: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """新增设备"""
    # 检查设备编码唯一性
    existing = await db.execute(
        select(Device).where(Device.device_code == req.device_code)
    )
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="设备编码已存在")

    device = Device(
        device_code=req.device_code,
        device_name=req.device_name,
        device_type=req.device_type,
        hierarchy_id=req.hierarchy_id,
        comm_mode=req.comm_mode,
        protocol_id=req.protocol_id,
        sim_card=req.sim_card,
        install_date=req.install_date,
        remark=req.remark,
    )
    db.add(device)
    await db.flush()
    await db.refresh(device)
    logger.info(f"新增设备: {device.device_code} (ID={device.id})")

    return ResponseModel(data={"id": device.id}, message="创建成功")


# ============ 修改设备 ============

@router.put("/{device_id}", response_model=ResponseModel)
async def update_device(
    device_id: int,
    req: DeviceUpdate,
    current_user: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """修改设备"""
    device = await db.get(Device, device_id)
    if not device:
        raise HTTPException(status_code=404, detail="设备不存在")

    update_data = req.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(device, key, value)

    await db.commit()
    logger.info(f"修改设备: ID={device_id}")
    return ResponseModel(message="修改成功")


# ============ 删除设备 ============

@router.delete("/{device_id}", response_model=ResponseModel)
async def delete_device(
    device_id: int,
    current_user: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """删除设备"""
    device = await db.get(Device, device_id)
    if not device:
        raise HTTPException(status_code=404, detail="设备不存在")

    # 检查是否关联数据点
    dp_result = await db.execute(
        select(func.count(DataPoint.id)).where(DataPoint.device_id == device_id)
    )
    dp_count = dp_result.scalar()
    if dp_count > 0:
        raise HTTPException(
            status_code=400,
            detail=f"该设备下有 {dp_count} 个数据点，请先解除关联",
        )

    await db.delete(device)
    await db.commit()
    logger.info(f"删除设备: ID={device_id}")
    return ResponseModel(message="删除成功")


# ============ 设备详情 ============

@router.get("/{device_id}", response_model=ResponseModel)
async def get_device(
    device_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """设备详情"""
    device = await db.get(Device, device_id)
    if not device:
        raise HTTPException(status_code=404, detail="设备不存在")

    return ResponseModel(data={
        "id": device.id,
        "device_code": device.device_code,
        "device_name": device.device_name,
        "device_type": device.device_type,
        "hierarchy_id": device.hierarchy_id,
        "comm_mode": device.comm_mode,
        "protocol_id": device.protocol_id,
        "sim_card": device.sim_card,
        "install_date": str(device.install_date) if device.install_date else None,
        "status": device.status,
        "last_heartbeat": device.last_heartbeat.isoformat() if device.last_heartbeat else None,
        "firmware_version": device.firmware_version,
        "remark": device.remark,
        "created_at": device.created_at.isoformat() if device.created_at else None,
        "updated_at": device.updated_at.isoformat() if device.updated_at else None,
    })


# ============ 设备状态查询 ============

@router.get("/{device_id}/status", response_model=ResponseModel)
async def get_device_status(
    device_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """设备状态/心跳信息"""
    device = await db.get(Device, device_id)
    if not device:
        raise HTTPException(status_code=404, detail="设备不存在")

    return ResponseModel(data={
        "device_id": device.id,
        "device_code": device.device_code,
        "status": device.status,
        "last_heartbeat": device.last_heartbeat.isoformat() if device.last_heartbeat else None,
        "firmware_version": device.firmware_version,
        "comm_mode": device.comm_mode,
    })


# ============ 批量导入 ============

@router.post("/import", response_model=ResponseModel)
async def import_devices(
    file: UploadFile = File(...),
    current_user: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """从Excel批量导入设备"""
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="请上传Excel文件(.xlsx)")

    content = await file.read()
    wb = load_workbook(BytesIO(content), read_only=True)
    ws = wb.active

    # 表头映射
    headers = [cell.value for cell in ws[1]]
    header_map = {}
    for idx, h in enumerate(headers):
        if h:
            header_map[h.strip()] = idx

    required_fields = ["设备编码"]
    for field in required_fields:
        if field not in header_map:
            raise HTTPException(status_code=400, detail=f"缺少必填列: {field}")

    imported = 0
    skipped = 0
    errors = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            device_code = str(row[header_map["设备编码"]]).strip() if row[header_map["设备编码"]] else None
            if not device_code:
                continue

            # 检查是否已存在
            existing = await db.execute(
                select(Device).where(Device.device_code == device_code)
            )
            if existing.scalar_one_or_none():
                skipped += 1
                continue

            device = Device(
                device_code=device_code,
                device_name=str(row[header_map.get("设备名称", -1)]).strip() if "设备名称" in header_map and row[header_map["设备名称"]] else None,
                device_type=str(row[header_map.get("设备类型", -1)]).strip() if "设备类型" in header_map and row[header_map["设备类型"]] else "4g_dtu",
                hierarchy_id=int(row[header_map["所属层级ID"]]) if "所属层级ID" in header_map and row[header_map["所属层级ID"]] else None,
                comm_mode=str(row[header_map.get("通讯方式", -1)]).strip() if "通讯方式" in header_map and row[header_map["通讯方式"]] else "mqtt",
                protocol_id=int(row[header_map["协议模板ID"]]) if "协议模板ID" in header_map and row[header_map["协议模板ID"]] else None,
                sim_card=str(row[header_map.get("SIM卡号", -1)]).strip() if "SIM卡号" in header_map and row[header_map["SIM卡号"]] else None,
                remark=str(row[header_map.get("备注", -1)]).strip() if "备注" in header_map and row[header_map["备注"]] else None,
            )
            db.add(device)
            imported += 1
        except Exception as e:
            errors.append(f"第{row_idx}行: {str(e)}")

    await db.commit()
    wb.close()

    return ResponseModel(data={
        "imported": imported,
        "skipped": skipped,
        "errors": errors[:50],  # 最多返回50条错误
    }, message=f"导入完成: 新增 {imported}, 跳过 {skipped}")


# ============ 导出Excel ============

@router.get("/export/excel")
async def export_devices(
    hierarchy_id: Optional[int] = None,
    status: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """导出设备列表到Excel"""
    query = select(Device)
    if hierarchy_id is not None:
        query = query.where(Device.hierarchy_id == hierarchy_id)
    if status:
        query = query.where(Device.status == status)
    query = query.order_by(Device.id)
    result = await db.execute(query)
    devices = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "设备列表"

    # 表头
    headers = [
        "ID", "设备编码", "设备名称", "设备类型", "所属层级ID",
        "通讯方式", "协议模板ID", "SIM卡号", "安装日期", "状态",
        "最后心跳", "固件版本", "备注", "创建时间",
    ]
    ws.append(headers)

    # 数据行
    for d in devices:
        ws.append([
            d.id,
            d.device_code,
            d.device_name,
            d.device_type,
            d.hierarchy_id,
            d.comm_mode,
            d.protocol_id,
            d.sim_card,
            str(d.install_date) if d.install_date else "",
            d.status,
            d.last_heartbeat.isoformat() if d.last_heartbeat else "",
            d.firmware_version,
            d.remark,
            d.created_at.isoformat() if d.created_at else "",
        ])

    # 设置列宽
    for col_idx, col in enumerate(ws.columns, 1):
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "A"].width = min(max_length + 4, 40)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    wb.close()

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=devices.xlsx"},
    )
